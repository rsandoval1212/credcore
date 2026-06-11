"""
Exportaciones a Excel para todos los módulos de CredCore.
Genera archivos .xlsx con formato profesional usando openpyxl.
"""
import io
from datetime import date, timedelta
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.throttling import UserRateThrottle


# FIX A5: Rate limit para exports (previene abuso sin bloquear uso normal)
class ExportThrottle(UserRateThrottle):
    rate = '30/minute'


# ── Helpers de formato ────────────────────────────────────────────────────────

def _excel_response(wb, filename: str) -> HttpResponse:
    """Devuelve un HttpResponse con el workbook como descarga."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_header(ws, row: int, cols: list[str], colors=None):
    """Aplica estilo de encabezado a una fila."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    DARK  = '1E3A5F'
    fill  = PatternFill('solid', fgColor=colors or DARK)
    font  = Font(bold=True, color='FFFFFF', size=10)
    align = Alignment(horizontal='center', vertical='center')
    thin  = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border


def _style_data_row(ws, row: int, num_cols: int, even: bool):
    from openpyxl.styles import PatternFill, Alignment, Border, Side
    fill = PatternFill('solid', fgColor='F9FAFB' if even else 'FFFFFF')
    thin = Side(border_style='thin', color='E5E7EB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = border
        if cell.alignment.horizontal is None:
            cell.alignment = Alignment(vertical='center')


def _fmt_decimal(v) -> float:
    """Convierte Decimal/None a float."""
    if v is None:
        return 0.0
    return float(v)


def _title_block(ws, title: str, subtitle: str):
    """Bloque de título en las primeras filas."""
    from openpyxl.styles import Font, Alignment, PatternFill
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16, color='1E3A5F')
    ws['A2'] = subtitle
    ws['A2'].font = Font(size=10, color='6B7280', italic=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18


# ── Exportar Clientes ─────────────────────────────────────────────────────────

class ExportCustomersView(APIView):
    permission_classes = [IsAuthenticated]
    throttle_classes = [ExportThrottle]

    def get(self, request):
        from apps.customers.models import Customer
        import openpyxl
        from openpyxl.styles import Font, Alignment, numbers

        qs = Customer.objects.filter(is_deleted=False).select_related('branch').order_by('last_name')
        today = date.today().strftime('%Y-%m-%d')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clientes'

        _title_block(ws, 'Cartera de Clientes — CredCore', f'Generado: {today} | Total: {qs.count()} clientes')

        headers = [
            'Código', 'Tipo', 'Nombre Completo', 'Cédula/RNC',
            'Teléfono', 'Email', 'Sucursal', 'Estado', 'Riesgo',
            'Score Crédito', 'Ingresos Mensuales', 'Préstamos Activos',
            'Saldo Pendiente', 'Registro',
        ]
        _style_header(ws, 4, headers)

        for i, c in enumerate(qs, 1):
            r = i + 4
            ws.cell(r, 1, c.customer_code)
            ws.cell(r, 2, 'Natural' if c.customer_type == 'NATURAL' else 'Jurídica')
            ws.cell(r, 3, c.get_full_name())
            ws.cell(r, 4, c.id_number)
            ws.cell(r, 5, c.phone1)
            ws.cell(r, 6, c.email or '')
            ws.cell(r, 7, c.branch.name if c.branch else '')
            ws.cell(r, 8, c.get_status_display() if hasattr(c, 'get_status_display') else c.status)
            ws.cell(r, 9, c.get_risk_level_display() if hasattr(c, 'get_risk_level_display') else c.risk_level or '')
            ws.cell(r, 10, c.credit_score or 0)
            ws.cell(r, 11, _fmt_decimal(c.monthly_income)).number_format = '#,##0.00'
            ws.cell(r, 12, c.active_loans_count if hasattr(c, 'active_loans_count') else 0)
            ws.cell(r, 13, _fmt_decimal(c.outstanding_balance)).number_format = '#,##0.00'
            ws.cell(r, 14, c.created_at.strftime('%d/%m/%Y') if c.created_at else '')
            _style_data_row(ws, r, len(headers), i % 2 == 0)

        # Anchos de columna
        widths = [12, 10, 30, 16, 14, 25, 18, 10, 10, 10, 18, 12, 18, 12]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(4, col).column_letter].width = w

        ws.freeze_panes = 'A5'
        return _excel_response(wb, f'clientes_{today}.xlsx')


# ── Exportar Préstamos ────────────────────────────────────────────────────────

class ExportLoansView(APIView):
    permission_classes = [IsAuthenticated]
    throttle_classes = [ExportThrottle]

    def get(self, request):
        from apps.loans.models import Loan
        import openpyxl

        qs = Loan.objects.filter(is_deleted=False).select_related(
            'customer', 'product', 'branch', 'officer'
        ).order_by('-created_at')

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        today = date.today().strftime('%Y-%m-%d')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Préstamos'

        _title_block(ws, 'Cartera de Préstamos — CredCore', f'Generado: {today} | Total: {qs.count()} préstamos')

        headers = [
            'N° Préstamo', 'Cliente', 'Producto', 'Sucursal', 'Oficial',
            'Capital Original', 'Tasa Anual %', 'Plazo (meses)',
            'Cuota Mensual', 'Saldo Capital', 'Saldo Interés', 'Mora',
            'Total Pendiente', 'Total Pagado', 'Estado',
            'Días Mora', 'Cuotas Pagadas', 'Cuotas Rest.',
            'Desembolso', 'Vencimiento',
        ]
        _style_header(ws, 4, headers)

        for i, loan in enumerate(qs, 1):
            r = i + 4
            total_out = _fmt_decimal(loan.outstanding_principal) + _fmt_decimal(loan.outstanding_interest) + _fmt_decimal(loan.outstanding_late_fees)
            ws.cell(r, 1, loan.loan_number)
            ws.cell(r, 2, loan.customer.get_full_name() if loan.customer else '')
            ws.cell(r, 3, loan.product.name if loan.product else '')
            ws.cell(r, 4, loan.branch.name if loan.branch else '')
            ws.cell(r, 5, loan.officer.get_full_name() if loan.officer else '')
            ws.cell(r, 6, _fmt_decimal(loan.principal_amount)).number_format = '#,##0.00'
            ws.cell(r, 7, _fmt_decimal(loan.annual_interest_rate))
            ws.cell(r, 8, loan.term_months)
            ws.cell(r, 9, _fmt_decimal(loan.monthly_payment)).number_format = '#,##0.00'
            ws.cell(r, 10, _fmt_decimal(loan.outstanding_principal)).number_format = '#,##0.00'
            ws.cell(r, 11, _fmt_decimal(loan.outstanding_interest)).number_format = '#,##0.00'
            ws.cell(r, 12, _fmt_decimal(loan.outstanding_late_fees)).number_format = '#,##0.00'
            ws.cell(r, 13, total_out).number_format = '#,##0.00'
            ws.cell(r, 14, _fmt_decimal(loan.total_paid)).number_format = '#,##0.00'
            ws.cell(r, 15, loan.get_status_display())
            ws.cell(r, 16, loan.days_past_due)
            ws.cell(r, 17, loan.installments_paid)
            ws.cell(r, 18, loan.installments_remaining)
            ws.cell(r, 19, loan.disbursement_date.strftime('%d/%m/%Y') if loan.disbursement_date else '')
            ws.cell(r, 20, loan.maturity_date.strftime('%d/%m/%Y') if loan.maturity_date else '')
            _style_data_row(ws, r, len(headers), i % 2 == 0)

        widths = [16, 28, 20, 16, 20, 16, 10, 12, 14, 14, 14, 12, 14, 14, 12, 10, 12, 12, 14, 14]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(4, col).column_letter].width = w
        ws.freeze_panes = 'A5'
        return _excel_response(wb, f'prestamos_{today}.xlsx')


# ── Exportar Cobros ───────────────────────────────────────────────────────────

class ExportPaymentsView(APIView):
    permission_classes = [IsAuthenticated]
    throttle_classes = [ExportThrottle]

    def get(self, request):
        from apps.payments.models import Payment
        import openpyxl

        qs = Payment.objects.filter(is_deleted=False).select_related(
            'loan', 'customer', 'received_by'
        ).order_by('-payment_date')

        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        if date_from: qs = qs.filter(payment_date__gte=date_from)
        if date_to:   qs = qs.filter(payment_date__lte=date_to)

        today = date.today().strftime('%Y-%m-%d')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cobros'

        _title_block(ws, 'Registro de Cobros — CredCore', f'Generado: {today} | Total: {qs.count()} pagos')

        headers = [
            'N° Recibo', 'N° Pago', 'Préstamo', 'Cliente',
            'Fecha', 'Total Cobrado', 'Capital', 'Interés', 'Mora',
            'Tipo', 'Método', 'Estado', 'Recibido por',
            'Banco', 'Referencia',
        ]
        _style_header(ws, 4, headers)

        for i, p in enumerate(qs, 1):
            r = i + 4
            ws.cell(r, 1, p.receipt_number)
            ws.cell(r, 2, p.payment_number)
            ws.cell(r, 3, p.loan.loan_number if p.loan else '')
            ws.cell(r, 4, p.customer.get_full_name() if p.customer else '')
            ws.cell(r, 5, p.payment_date.strftime('%d/%m/%Y') if p.payment_date else '')
            ws.cell(r, 6, _fmt_decimal(p.total_amount)).number_format = '#,##0.00'
            ws.cell(r, 7, _fmt_decimal(p.principal_amount)).number_format = '#,##0.00'
            ws.cell(r, 8, _fmt_decimal(p.interest_amount)).number_format = '#,##0.00'
            ws.cell(r, 9, _fmt_decimal(p.late_fee_amount)).number_format = '#,##0.00'
            ws.cell(r, 10, p.get_payment_type_display())
            ws.cell(r, 11, p.get_payment_method_display())
            ws.cell(r, 12, p.get_status_display())
            ws.cell(r, 13, p.received_by.get_full_name() if p.received_by else '')
            ws.cell(r, 14, p.bank_name or '')
            ws.cell(r, 15, p.reference_number or '')
            _style_data_row(ws, r, len(headers), i % 2 == 0)

        widths = [14, 14, 16, 28, 12, 14, 14, 12, 12, 18, 16, 12, 20, 18, 16]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(4, col).column_letter].width = w
        ws.freeze_panes = 'A5'
        return _excel_response(wb, f'cobros_{today}.xlsx')


# ── Reporte Maestro (todas las hojas + estadísticas) ─────────────────────────

class ExportMasterReportView(APIView):
    """
    Genera un archivo Excel completo con múltiples hojas:
    - Resumen ejecutivo con KPIs
    - Cartera de préstamos
    - Clientes activos
    - Cobros del período
    - Análisis de mora
    - Estadísticas de ingresos
    Se actualiza con los datos en tiempo real cada vez que se descarga.
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [ExportThrottle]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter

        from apps.loans.models import Loan
        from apps.payments.models import Payment
        from apps.customers.models import Customer
        from django.db.models.functions import TruncMonth

        today = date.today()
        today_str = today.strftime('%d/%m/%Y')
        month_start = today.replace(day=1)
        year_start  = today.replace(month=1, day=1)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # quitar hoja por defecto

        # ── Paleta ──────────────────────────────────────────────────────────
        DARK_BLUE  = '1E3A5F'
        MED_BLUE   = '2563EB'
        LIGHT_BLUE = 'DBEAFE'
        GREEN      = '059669'
        AMBER      = 'D97706'
        RED        = 'DC2626'
        GRAY       = 'F9FAFB'

        def cell_style(cell, bold=False, size=10, color='000000', bg=None, align='left', number_fmt=None):
            cell.font = Font(bold=bold, size=size, color=color)
            if bg:
                cell.fill = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            if number_fmt:
                cell.number_format = number_fmt

        def header_row(ws, row, cols, bg=DARK_BLUE):
            thin = Side(border_style='thin', color='CCCCCC')
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
            for i, col in enumerate(cols, 1):
                c = ws.cell(row=row, column=i, value=col)
                c.font      = Font(bold=True, size=10, color='FFFFFF')
                c.fill      = PatternFill('solid', fgColor=bg)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = bdr
            ws.row_dimensions[row].height = 20

        # ════════════════════════════════════════════════════════════════════
        # HOJA 1: RESUMEN EJECUTIVO
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.create_sheet('📊 Resumen Ejecutivo')

        # Título
        ws1.merge_cells('A1:H1')
        ws1['A1'] = '  REPORTE EJECUTIVO — CREDCORE'
        cell_style(ws1['A1'], bold=True, size=18, color='FFFFFF', bg=DARK_BLUE, align='left')
        ws1.row_dimensions[1].height = 36

        ws1.merge_cells('A2:H2')
        ws1['A2'] = f'  Generado el {today_str}  |  Datos en tiempo real'
        cell_style(ws1['A2'], size=10, color='6B7280', bg='F3F4F6', align='left')
        ws1.row_dimensions[2].height = 18

        # Métricas de préstamos
        loan_qs  = Loan.objects.filter(is_deleted=False)
        pay_qs   = Payment.objects.filter(status='CONFIRMED')
        cust_qs  = Customer.objects.filter(is_deleted=False)

        active_loans = loan_qs.filter(status='ACTIVE')
        overdue      = active_loans.filter(days_past_due__gt=0)
        agg_loans    = active_loans.aggregate(
            portfolio=Sum('outstanding_principal'),
            overdue_portfolio=Sum('outstanding_principal', filter=Q(days_past_due__gt=0)),
        )
        agg_pay_month = pay_qs.filter(payment_date__gte=month_start).aggregate(
            total=Sum('total_amount'), interest=Sum('interest_amount'), count=Count('id')
        )
        agg_pay_year = pay_qs.filter(payment_date__gte=year_start).aggregate(
            total=Sum('total_amount'), interest=Sum('interest_amount')
        )

        portfolio      = _fmt_decimal(agg_loans['portfolio'])
        overdue_pf     = _fmt_decimal(agg_loans['overdue_portfolio'])
        delinq_rate    = round((overdue_pf / portfolio * 100), 2) if portfolio > 0 else 0
        month_income   = _fmt_decimal(agg_pay_month['interest'])
        month_cobros   = _fmt_decimal(agg_pay_month['total'])
        year_income    = _fmt_decimal(agg_pay_year['interest'])

        kpis = [
            ('PORTAFOLIO ACTIVO',       portfolio,    MED_BLUE,  '#,##0.00'),
            ('CARTERA EN MORA',         overdue_pf,   RED,       '#,##0.00'),
            ('TASA MOROSIDAD',          delinq_rate,  AMBER,     '0.00"%"'),
            ('CLIENTES ACTIVOS',        cust_qs.filter(status='ACTIVE').count(), GREEN, '0'),
            ('COBROS DEL MES',          month_cobros, MED_BLUE,  '#,##0.00'),
            ('INTERESES DEL MES',       month_income, GREEN,     '#,##0.00'),
            ('INTERESES DEL AÑO',       year_income,  GREEN,     '#,##0.00'),
            ('PRÉSTAMOS ACTIVOS',       active_loans.count(), MED_BLUE, '0'),
        ]

        ws1.row_dimensions[4].height = 14
        ws1.row_dimensions[5].height = 52
        ws1.row_dimensions[6].height = 16
        ws1.row_dimensions[7].height = 22

        for col, (label, value, color, fmt) in enumerate(kpis, 1):
            lc = ws1.cell(5, col, f' {label}')
            cell_style(lc, bold=True, size=8, color='6B7280', bg='F9FAFB', align='left')
            vc = ws1.cell(6, col, value)
            cell_style(vc, bold=True, size=16, color=color, bg='FFFFFF', align='center', number_fmt=fmt)
            ws1.column_dimensions[get_column_letter(col)].width = 20

        # Tabla por estado de préstamos
        ws1.row_dimensions[9].height = 14
        ws1.merge_cells('A9:D9')
        ws1['A9'] = '  PRÉSTAMOS POR ESTADO'
        cell_style(ws1['A9'], bold=True, size=11, color=DARK_BLUE, bg=LIGHT_BLUE)

        header_row(ws1, 10, ['Estado', 'Cantidad', 'Portafolio (RD$)', '% del Total'], MED_BLUE)
        by_status = loan_qs.values('status').annotate(
            count=Count('id'), total=Sum('outstanding_principal')
        ).order_by('-total')
        STATUS_LABELS = {
            'ACTIVE': 'Activo', 'COMPLETED': 'Completado', 'DEFAULTED': 'En Mora',
            'WRITTEN_OFF': 'Castigado', 'CANCELLED': 'Cancelado', 'REFINANCED': 'Refinanciado',
        }
        grand_total = loan_qs.aggregate(t=Sum('outstanding_principal'))['t'] or 1
        for i, row_data in enumerate(by_status, 11):
            pct = round(_fmt_decimal(row_data['total']) / _fmt_decimal(grand_total) * 100, 1)
            bg = 'F9FAFB' if i % 2 == 0 else 'FFFFFF'
            for col, val in enumerate([
                STATUS_LABELS.get(row_data['status'], row_data['status']),
                row_data['count'],
                _fmt_decimal(row_data['total']),
                pct,
            ], 1):
                c = ws1.cell(i, col, val)
                c.fill = PatternFill('solid', fgColor=bg)
                c.alignment = Alignment(horizontal='right' if col > 1 else 'left', vertical='center')
                if col == 3: c.number_format = '#,##0.00'
                if col == 4: c.number_format = '0.0"%"'

        # ════════════════════════════════════════════════════════════════════
        # HOJA 2: CARTERA DE PRÉSTAMOS
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('💰 Préstamos')
        ws2.merge_cells('A1:R1')
        ws2['A1'] = f'CARTERA DE PRÉSTAMOS — Generado {today_str}'
        cell_style(ws2['A1'], bold=True, size=13, color='FFFFFF', bg=DARK_BLUE)
        ws2.row_dimensions[1].height = 24

        headers2 = ['N° Préstamo', 'Cliente', 'Producto', 'Sucursal',
                    'Capital', 'Tasa %', 'Plazo', 'Cuota',
                    'Saldo Cap.', 'Saldo Int.', 'Mora', 'Total Pend.',
                    'Estado', 'Días Mora', 'Desembolso', 'Vencimiento']
        header_row(ws2, 3, headers2)

        # FIX #15: Limitar a 10000 registros para evitar problemas de memoria
        all_loans = loan_qs.select_related('customer', 'product', 'branch').order_by('-created_at')[:10000]
        for i, ln in enumerate(all_loans, 4):
            tot = _fmt_decimal(ln.outstanding_principal) + _fmt_decimal(ln.outstanding_interest) + _fmt_decimal(ln.outstanding_late_fees)
            bg = 'FFF7ED' if ln.days_past_due > 30 else ('FEF3C7' if ln.days_past_due > 0 else ('F9FAFB' if i % 2 == 0 else 'FFFFFF'))
            vals = [
                ln.loan_number, ln.customer.get_full_name() if ln.customer else '',
                ln.product.name if ln.product else '', ln.branch.name if ln.branch else '',
                _fmt_decimal(ln.principal_amount), _fmt_decimal(ln.annual_interest_rate),
                ln.term_months, _fmt_decimal(ln.monthly_payment),
                _fmt_decimal(ln.outstanding_principal), _fmt_decimal(ln.outstanding_interest),
                _fmt_decimal(ln.outstanding_late_fees), tot,
                ln.get_status_display(), ln.days_past_due,
                ln.disbursement_date.strftime('%d/%m/%Y') if ln.disbursement_date else '',
                ln.maturity_date.strftime('%d/%m/%Y') if ln.maturity_date else '',
            ]
            money_cols = {5, 8, 9, 10, 11, 12}
            for col, val in enumerate(vals, 1):
                c = ws2.cell(i, col, val)
                c.fill = PatternFill('solid', fgColor=bg)
                if col in money_cols: c.number_format = '#,##0.00'

        col_widths2 = [16, 28, 20, 16, 14, 8, 8, 14, 14, 14, 12, 14, 12, 10, 14, 14]
        for col, w in enumerate(col_widths2, 1):
            ws2.column_dimensions[get_column_letter(col)].width = w
        ws2.freeze_panes = 'A4'

        # ════════════════════════════════════════════════════════════════════
        # HOJA 3: CLIENTES
        # ════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('👥 Clientes')
        ws3.merge_cells('A1:M1')
        ws3['A1'] = f'CARTERA DE CLIENTES — Generado {today_str}'
        cell_style(ws3['A1'], bold=True, size=13, color='FFFFFF', bg=DARK_BLUE)
        ws3.row_dimensions[1].height = 24

        headers3 = ['Código', 'Nombre', 'Cédula', 'Teléfono', 'Email',
                    'Sucursal', 'Estado', 'Riesgo', 'Score',
                    'Ingresos', 'P. Activos', 'Saldo', 'Registro']
        header_row(ws3, 3, headers3)

        for i, c in enumerate(cust_qs.select_related('branch').order_by('last_name'), 4):
            bg = 'F9FAFB' if i % 2 == 0 else 'FFFFFF'
            vals = [
                c.customer_code, c.get_full_name(), c.id_number, c.phone1, c.email or '',
                c.branch.name if c.branch else '',
                c.get_status_display() if hasattr(c, 'get_status_display') else c.status,
                c.get_risk_level_display() if hasattr(c, 'get_risk_level_display') else (c.risk_level or ''),
                c.credit_score or 0, _fmt_decimal(c.monthly_income),
                0, _fmt_decimal(c.outstanding_balance if hasattr(c, 'outstanding_balance') else 0),
                c.created_at.strftime('%d/%m/%Y') if c.created_at else '',
            ]
            for col, val in enumerate(vals, 1):
                cv = ws3.cell(i, col, val)
                cv.fill = PatternFill('solid', fgColor=bg)
                if col in {10, 12}: cv.number_format = '#,##0.00'

        col_widths3 = [12, 28, 16, 14, 24, 16, 10, 10, 8, 16, 10, 14, 12]
        for col, w in enumerate(col_widths3, 1):
            ws3.column_dimensions[get_column_letter(col)].width = w
        ws3.freeze_panes = 'A4'

        # ════════════════════════════════════════════════════════════════════
        # HOJA 4: COBROS
        # ════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet('💵 Cobros')
        ws4.merge_cells('A1:N1')
        ws4['A1'] = f'REGISTRO DE COBROS — Generado {today_str}'
        cell_style(ws4['A1'], bold=True, size=13, color='FFFFFF', bg=DARK_BLUE)
        ws4.row_dimensions[1].height = 24

        headers4 = ['Recibo', 'Préstamo', 'Cliente', 'Fecha',
                    'Total', 'Capital', 'Interés', 'Mora',
                    'Tipo', 'Método', 'Estado', 'Recibido por', 'Banco', 'Referencia']
        header_row(ws4, 3, headers4)

        recent_pays = pay_qs.select_related('loan', 'customer', 'received_by').order_by('-payment_date')[:5000]
        for i, p in enumerate(recent_pays, 4):
            bg = 'F9FAFB' if i % 2 == 0 else 'FFFFFF'
            vals = [
                p.receipt_number, p.loan.loan_number if p.loan else '',
                p.customer.get_full_name() if p.customer else '',
                p.payment_date.strftime('%d/%m/%Y') if p.payment_date else '',
                _fmt_decimal(p.total_amount), _fmt_decimal(p.principal_amount),
                _fmt_decimal(p.interest_amount), _fmt_decimal(p.late_fee_amount),
                p.get_payment_type_display(), p.get_payment_method_display(), p.get_status_display(),
                p.received_by.get_full_name() if p.received_by else '',
                p.bank_name or '', p.reference_number or '',
            ]
            for col, val in enumerate(vals, 1):
                cv = ws4.cell(i, col, val)
                cv.fill = PatternFill('solid', fgColor=bg)
                if col in {5, 6, 7, 8}: cv.number_format = '#,##0.00'

        col_widths4 = [14, 16, 28, 12, 14, 14, 12, 12, 18, 16, 12, 20, 16, 16]
        for col, w in enumerate(col_widths4, 1):
            ws4.column_dimensions[get_column_letter(col)].width = w
        ws4.freeze_panes = 'A4'

        # ════════════════════════════════════════════════════════════════════
        # HOJA 5: INGRESOS POR MES (últimos 12 meses)
        # ════════════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet('📈 Ingresos por Mes')
        ws5.merge_cells('A1:F1')
        ws5['A1'] = f'INGRESOS POR MES (Últimos 12 meses) — Generado {today_str}'
        cell_style(ws5['A1'], bold=True, size=13, color='FFFFFF', bg=DARK_BLUE)
        ws5.row_dimensions[1].height = 24

        header_row(ws5, 3, ['Mes', 'N° Cobros', 'Total Cobrado', 'Capital', 'Intereses', 'Mora'])

        monthly_pay = (
            Payment.objects.filter(status='CONFIRMED', payment_date__gte=year_start - timedelta(days=365))
            .annotate(month=TruncMonth('payment_date'))
            .values('month')
            .annotate(
                cnt=Count('id'),
                total=Sum('total_amount'),
                principal=Sum('principal_amount'),
                interest=Sum('interest_amount'),
                late_fees=Sum('late_fee_amount'),
            )
            .order_by('month')
        )

        MONTHS_ES = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        for i, m in enumerate(monthly_pay, 4):
            d = m['month'].date() if hasattr(m['month'], 'date') else m['month']
            label = f"{MONTHS_ES[d.month]} {d.year}"
            bg = 'F9FAFB' if i % 2 == 0 else 'FFFFFF'
            for col, val in enumerate([
                label, m['cnt'],
                _fmt_decimal(m['total']), _fmt_decimal(m['principal']),
                _fmt_decimal(m['interest']), _fmt_decimal(m['late_fees']),
            ], 1):
                c = ws5.cell(i, col, val)
                c.fill = PatternFill('solid', fgColor=bg)
                if col >= 3: c.number_format = '#,##0.00'

        # Gráfico de ingresos
        if monthly_pay.count() > 0:
            chart = BarChart()
            chart.type = 'col'
            chart.title = 'Cobros Mensuales'
            chart.y_axis.title = 'RD$'
            chart.x_axis.title = 'Mes'
            chart.style = 10
            chart.width  = 22
            chart.height = 12

            last_row = 3 + monthly_pay.count()
            data_ref  = Reference(ws5, min_col=3, max_col=5, min_row=3, max_row=last_row)
            cats_ref  = Reference(ws5, min_col=1, min_row=4, max_row=last_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws5.add_chart(chart, 'A' + str(last_row + 3))

        col_widths5 = [14, 10, 16, 16, 16, 14]
        for col, w in enumerate(col_widths5, 1):
            ws5.column_dimensions[get_column_letter(col)].width = w

        # ════════════════════════════════════════════════════════════════════
        # HOJA 6: ANÁLISIS DE MORA
        # ════════════════════════════════════════════════════════════════════
        ws6 = wb.create_sheet('⚠️ Análisis de Mora')
        ws6.merge_cells('A1:F1')
        ws6['A1'] = f'ANÁLISIS DE MORA — Generado {today_str}'
        cell_style(ws6['A1'], bold=True, size=13, color='FFFFFF', bg=RED)
        ws6.row_dimensions[1].height = 24

        header_row(ws6, 3, ['Rango de Mora', 'Préstamos', 'Portafolio', '% Portafolio', 'Clientes', 'Promedio/Préstamo'], RED)

        buckets = [
            ('Al día (0 días)',      0, 0),
            ('1 - 30 días',         1, 30),
            ('31 - 60 días',        31, 60),
            ('61 - 90 días',        61, 90),
            ('91 - 120 días',       91, 120),
            ('Más de 120 días',     121, 9999),
        ]
        total_port = _fmt_decimal(active_loans.aggregate(t=Sum('outstanding_principal'))['t'])
        for i, (label, mn, mx) in enumerate(buckets, 4):
            qs_b = active_loans
            if mn == 0 and mx == 0:
                qs_b = qs_b.filter(days_past_due=0)
            else:
                qs_b = qs_b.filter(days_past_due__gte=mn, days_past_due__lte=mx)
            agg_b = qs_b.aggregate(cnt=Count('id'), total=Sum('outstanding_principal'), clients=Count('customer', distinct=True))
            cnt  = agg_b['cnt'] or 0
            tot  = _fmt_decimal(agg_b['total'])
            pct  = round(tot / total_port * 100, 2) if total_port > 0 else 0
            avg  = round(tot / cnt, 2) if cnt > 0 else 0
            bg   = 'FEF2F2' if mn >= 31 else ('FFFBEB' if mn >= 1 else 'F0FDF4')
            for col, val in enumerate([label, cnt, tot, pct, agg_b['clients'] or 0, avg], 1):
                c = ws6.cell(i, col, val)
                c.fill = PatternFill('solid', fgColor=bg)
                if col in {3, 6}: c.number_format = '#,##0.00'
                if col == 4: c.number_format = '0.00"%"'

        col_widths6 = [22, 12, 18, 14, 12, 18]
        for col, w in enumerate(col_widths6, 1):
            ws6.column_dimensions[get_column_letter(col)].width = w

        # Gráfico de mora (torta)
        pie = PieChart()
        pie.title = 'Distribución de Mora'
        pie.style = 10
        pie.width  = 18
        pie.height = 12
        data_ref_pie = Reference(ws6, min_col=3, min_row=3, max_row=9)
        cats_ref_pie = Reference(ws6, min_col=1, min_row=4, max_row=9)
        pie.add_data(data_ref_pie, titles_from_data=True)
        pie.set_categories(cats_ref_pie)
        ws6.add_chart(pie, 'H4')

        # ════════════════════════════════════════════════════════════════════
        # Propiedades del workbook
        # ════════════════════════════════════════════════════════════════════
        wb.properties.title    = 'Reporte CredCore'
        wb.properties.subject  = f'Reporte generado el {today_str}'
        wb.properties.creator  = 'CredCore Sistema de Créditos'

        filename = f'credcore_reporte_maestro_{today.strftime("%Y%m%d")}.xlsx'
        return _excel_response(wb, filename)
