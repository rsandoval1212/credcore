"""
Importación masiva desde Excel — CredCore
Permite cargar clientes y préstamos desde archivos .xlsx
"""
import io
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser


class ImportTemplateView(APIView):
    """Descarga la plantilla Excel para importar clientes."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clientes'

        headers = [
            'primer_nombre*', 'segundo_nombre', 'primer_apellido*', 'segundo_apellido',
            'cedula*', 'fecha_nacimiento (DD/MM/YYYY)', 'sexo (M/F)',
            'telefono*', 'telefono2', 'email', 'whatsapp',
            'direccion', 'sector', 'municipio', 'provincia',
            'ingreso_mensual', 'gastos_mensuales', 'ocupacion',
            'sucursal_id (opcional)',
        ]

        BLUE = 'FF1E3A5F'
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.font      = Font(bold=True, color='FFFFFFFF', size=9)
            c.fill      = PatternFill('solid', fgColor=BLUE)
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[ws.cell(1, i).column_letter].width = max(len(h) + 4, 14)

        # Fila de ejemplo
        example = [
            'Juan', 'Carlos', 'Pérez', 'Martínez',
            '001-0000001-1', '15/06/1985', 'M',
            '809-555-0001', '809-555-0002', 'juan@email.com', '8095550001',
            'Calle Principal #5', 'Los Prados', 'Santo Domingo Norte', 'Santo Domingo',
            '35000', '15000', 'Empleado',
            '1',
        ]
        for i, v in enumerate(example, 1):
            ws.cell(2, i, v)

        # Hoja de instrucciones
        ws2 = wb.create_sheet('Instrucciones')
        instrucciones = [
            ('Columna', 'Descripción', 'Requerido', 'Ejemplo'),
            ('primer_nombre', 'Primer nombre del cliente', 'Sí', 'Juan'),
            ('cedula', 'Cédula o documento de identidad', 'Sí', '001-0000001-1'),
            ('fecha_nacimiento', 'Formato DD/MM/YYYY', 'No', '15/06/1985'),
            ('sexo', 'M para masculino, F para femenino', 'No', 'M'),
            ('telefono', 'Teléfono principal', 'Sí', '809-555-0001'),
            ('ingreso_mensual', 'Ingreso mensual en RD$', 'No', '35000'),
            ('sucursal_id', 'ID de sucursal — dejar vacío usa la sucursal principal', 'No', '1'),
        ]
        for i, row in enumerate(instrucciones, 1):
            for j, val in enumerate(row, 1):
                c = ws2.cell(i, j, val)
                if i == 1:
                    c.font = Font(bold=True)

        # Hoja de sucursales disponibles
        ws3 = wb.create_sheet('Sucursales')
        from apps.branches.models import Branch
        ws3.cell(1, 1, 'ID').font = Font(bold=True)
        ws3.cell(1, 2, 'Nombre').font = Font(bold=True)
        for i, b in enumerate(Branch.objects.filter(is_active=True), 2):
            ws3.cell(i, 1, b.id)
            ws3.cell(i, 2, b.name)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_importar_clientes.xlsx"'
        return response


class ImportCustomersView(APIView):
    """Importa clientes desde un archivo Excel."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        import openpyxl
        from datetime import datetime
        from apps.customers.models import Customer
        from apps.branches.models import Branch

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Debes subir un archivo .xlsx'}, status=400)

        if not file.name.endswith('.xlsx'):
            return Response({'detail': 'Solo se aceptan archivos .xlsx'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb['Clientes']
        except Exception as e:
            return Response({'detail': f'Error al leer el archivo: {e}'}, status=400)

        created = 0
        skipped = 0
        errors  = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            try:
                first_name = str(row[0] or '').strip()
                second_name = str(row[1] or '').strip()
                last_name   = str(row[2] or '').strip()
                second_last = str(row[3] or '').strip()
                id_number   = str(row[4] or '').strip()
                dob_raw     = row[5]
                sex         = str(row[6] or 'M').strip().upper()
                phone1      = str(row[7] or '').strip()
                phone2      = str(row[8] or '').strip()
                email       = str(row[9] or '').strip()
                whatsapp    = str(row[10] or '').strip()
                address     = str(row[11] or '').strip()
                sector      = str(row[12] or '').strip()
                municipality = str(row[13] or '').strip()
                province    = str(row[14] or '').strip()
                income      = float(row[15]) if row[15] else None
                expenses    = float(row[16]) if row[16] else None
                occupation  = str(row[17] or '').strip()
                branch_id   = int(row[18]) if row[18] else None

                if not first_name or not last_name:
                    errors.append(f'Fila {row_num}: nombre y apellido son requeridos')
                    skipped += 1
                    continue

                if not id_number:
                    errors.append(f'Fila {row_num}: cédula requerida')
                    skipped += 1
                    continue

                # Verificar duplicado
                if Customer.objects.filter(id_number=id_number, is_deleted=False).exists():
                    errors.append(f'Fila {row_num}: cédula {id_number} ya existe (omitida)')
                    skipped += 1
                    continue

                # Parsear fecha
                dob = None
                if dob_raw:
                    try:
                        if isinstance(dob_raw, str):
                            dob = datetime.strptime(dob_raw.strip(), '%d/%m/%Y').date()
                        elif hasattr(dob_raw, 'date'):
                            dob = dob_raw.date()
                        else:
                            dob = dob_raw
                    except Exception:
                        pass

                # Sucursal: usa la enviada o la principal automáticamente
                branch = None
                if branch_id:
                    try:
                        branch = Branch.objects.get(id=branch_id)
                    except Branch.DoesNotExist:
                        branch = Branch.get_main()
                else:
                    branch = Branch.get_main()

                from apps.core.utils import generate_code
                Customer.objects.create(
                    first_name=first_name,
                    second_name=second_name,
                    last_name=last_name,
                    second_last_name=second_last,
                    id_number=id_number,
                    id_type='CEDULA',
                    date_of_birth=dob,
                    gender=sex if sex in ('M', 'F') else 'M',
                    phone1=phone1,
                    phone2=phone2,
                    email=email,
                    whatsapp=whatsapp,
                    address=address,
                    sector=sector,
                    municipality=municipality,
                    province=province,
                    country='República Dominicana',
                    branch=branch,
                    monthly_income=income,
                    monthly_expenses=expenses,
                    customer_type='NATURAL',
                    status='ACTIVE',
                    created_by=request.user,
                )
                created += 1

            except Exception as e:
                errors.append(f'Fila {row_num}: {str(e)}')
                skipped += 1

        return Response({
            'success': True,
            'created': created,
            'skipped': skipped,
            'errors':  errors[:20],  # máx 20 errores mostrados
            'message': f'{created} clientes importados, {skipped} omitidos.',
        })
